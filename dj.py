import openpyxl
from youtube_search import YoutubeSearch
import yt_dlp

# Set the path to the XLSX spreadsheet
xlsx_file = 'songs.xlsx'

# Set the name of the sheet containing the video titles and artist names
sheet = 'songs'

# Load the XLSX spreadsheet
workbook = openpyxl.load_workbook(xlsx_file)

# Get the active sheet
worksheet = workbook[sheet]

# Find the indexes of the 'Song' and 'Artist' columns
for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(row)
    for i, col in enumerate(row):
        if col == 'Song':
            song_col_index = i + 1
        elif col == 'Artist':
            artist_col_index = i + 1


headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
print(headers)

print(f"Song column index: {song_col_index}")
print(f"Artist column index: {artist_col_index}")

# Make sure both columns were found
if song_col_index is None or artist_col_index is None:
    print("Error: could not find 'Song' and/or 'Artist' columns")
    exit()

# Set the flag for downloading all videos to False initially
download_all = False

# Prompt the user to choose whether to download all videos or not
choice = input("Do you want to download all videos? (Y/N)")

# Set the flag to True if the user enters 'Y' or 'y'
if choice.lower() == 'y':
    download_all = True

# Create a YouTube downloader
ydl_opts = {
    'format': 'bestaudio/best',
    'postprocessors': [{
        'key': 'FFmpegExtractAudio',
        'preferredcodec': 'mp3',
        'preferredquality': '192',
    }],
}

# Iterate through the rows of the worksheet
# Iterate through the rows of the worksheet
for row in worksheet.iter_rows(min_row=2):
    video_id = row[5].value
    if video_id:
        # Video ID is already present, skip the search and move on to the next row
        print(f"Video ID already present: {video_id}")
        continue

    song_title = row[song_col_index - 1].value
    artist_name = row[artist_col_index - 1].value

    # Combine the song title and artist name to form the search query
    search_query = f"{song_title} {artist_name} lyric video"

    print(f"Searching for: {search_query}")

    # Search for the video using the search query
    results = YoutubeSearch(search_query, max_results=1).to_dict()

    # Retrieve the video ID of the first search result
    video_id = results[0]['id']

    # Write the video ID to the sixth column of the current row
    worksheet.cell(row=row[0].row, column=6).value = video_id if video_id else ""

    if download_all or input('Download? Y or N:') == 'Y':
        print(f"Downloading video: {video_id}")

        # Download the video as an MP3
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([f"https://www.youtube.com/watch?v={video_id}"])


# Save the changes to the XLSX spreadsheet
workbook.save(xlsx_file)

print('Done')    